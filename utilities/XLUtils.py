import allure
import openpyxl
from utilities.baseClass import BaseClass


class XLUtils:

    @staticmethod
    @allure.step('Get Test Data for Test case ({test_ID})')
    def getTestData(test_ID):
        data_dict, flag = {}, 0
        workBookName = BaseClass.ROOT_PATH + r"\testData\testData.xlsx"
        workBook = openpyxl.load_workbook(workBookName, True)
        workSheet = workBook["TestSheet"]

        for iRow in range(1, workSheet.max_row + 1):  # get rows
            if workSheet.cell(row=iRow, column=1).value == test_ID:

                for iCol in range(2, workSheet.max_column + 1):  # starting from 2 since we dont want testcase ID
                    # print(workSheet.cell(row=iRow, column=iCol).value)
                    data_dict[workSheet.cell(row=1, column=iCol).value] = workSheet.cell(row=iRow, column=iCol).value
                flag = 1
                break

        if flag != 1:
            raise ValueError(f"No data found for TestCaseID: {test_ID}")

        return [data_dict][0]

# print(workSheet.cell(row=1, column=2).value)
# print(workSheet['A1'].value)
# write
# workSheet.cell(row=3, column=3).value = 'Bob'
# workBook.save(BaseClass.ROOT_PATH+r"\testData\testData.xlsx")
# workBook.close
